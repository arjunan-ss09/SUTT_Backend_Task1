import openpyxl
import json
wb= openpyxl.load_workbook(r"/Users/arjunanss09/Downloads/file.xlsx")
sheet= wb.active
data={}
columns_no= sheet.max_column
for i in range(1,columns_no):
    menu={}
    breakfast=[]
    lunch=[]
    dinner=[]
    for j in range(4,13):
        item=sheet.cell(row=j, column=i).value
        if "*" not in str(item) and item is not None:
            item_=" ".join(item.split())
            breakfast.append(item_)
    for j in range(15,23):
        item=sheet.cell(row=j, column=i).value
        if "*" not in str(item) and item is not None:
            item_=" ".join(item.split())
            lunch.append(item_)
    for j in range(25,32):
        item=sheet.cell(row=j, column=i).value
        if "*" not in str(item) and item is not None:
            item_=" ".join(item.split())
            dinner.append(item_)
    menu["BREAKFAST"]=breakfast
    menu["LUNCH"]=lunch
    menu["DINNER"]=dinner
    if i<10:
        data["2025-02-0"+str(i)]=menu
    else:
        data["2025-02-"+str(i)]=menu
f=open('//Users//arjunanss09//Downloads//mess_menu.json',"w")
json.dump(data,f,indent=2)

    

        
        
    
    
